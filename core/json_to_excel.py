import json
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import os
import re

# Konfigurasi folder
INPUT_FOLDER = "output json"
OUTPUT_FOLDER = "output excel"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def normalize_spaces(value):
    """Hilangkan spasi ganda dan pastikan string rapi."""
    if not value:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip()

def create_excel_from_json(json_file):
    with open(json_file, 'r', encoding='utf-8') as f:
        port_data = json.load(f)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Port Data"
    
    headers = ["Port", "Description", "Status", "Vlan", "Duplex", "Speed", "Type", "Mac add", "IP"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    row = 2
    for port, data in port_data.items():
        macs = "\n".join([mac.get('mac', '') for mac in data.get('mac_addresses', [])])
        ips = "\n".join(data.get('ip_addresses', []))


        ws.cell(row=row, column=1, value=normalize_spaces(port))
        ws.cell(row=row, column=2, value=normalize_spaces(data.get('description', '')))
        ws.cell(row=row, column=3, value=normalize_spaces(data.get('status', '')))
        ws.cell(row=row, column=4, value=normalize_spaces(data.get('vlan', '')))
        ws.cell(row=row, column=5, value=normalize_spaces(data.get('duplex', '')))
        ws.cell(row=row, column=6, value=normalize_spaces(data.get('speed', '')))
        ws.cell(row=row, column=7, value=normalize_spaces(data.get('type', '')))
        ws.cell(row=row, column=8, value=macs)
        ws.cell(row=row, column=8).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=9, value=ips)
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True)
        row += 1
    
    # Auto width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    
    output_file = os.path.splitext(os.path.basename(json_file))[0] + '.xlsx'
    output_path = os.path.join(OUTPUT_FOLDER, output_file)
    wb.save(output_path)
    print(f"Created {output_path}")

def process_all_json_files():
    json_files = [f for f in os.listdir(INPUT_FOLDER) if f.endswith('.json')]
    if not json_files:
        print(f"No JSON files found in {INPUT_FOLDER}")
        return
    print(f"Converting {len(json_files)} JSON files to Excel...")
    for json_file in json_files:
        try:
            create_excel_from_json(os.path.join(INPUT_FOLDER, json_file))
        except Exception as e:
            print(f"Error processing {json_file}: {str(e)}")
    print("\nAll conversions completed!")

if __name__ == "__main__":
    process_all_json_files()
